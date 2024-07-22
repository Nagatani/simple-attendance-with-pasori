import sqlite3
import openpyxl

import config

import argparse
parser = argparse.ArgumentParser()
parser.add_argument("section_id", help="第何回かをパラメータで指定してください。 0を指定すると全部の回を1つのブックにまとめて出力します。", type=int)
args = parser.parse_args()

con = sqlite3.connect(config.database_name)
cur = con.cursor()


def export():
  # WorkBookの作成
  wb = openpyxl.Workbook()
  sheet = wb.active
  wb.remove(sheet) # Sheet1の削除

  ws = wb.create_sheet() # 第何回かのシートを作成
  ws.title = str(args.section_id)

  cur.execute("select id, card_id, student_id, forgot_card, create_datetime, update_datetime from attendance where section_id = ? order by student_id", (args.section_id, ))

  ws.append(['ROWNUMBER', 'カードID', '学籍番号', 'カード忘れ', '登録日時', '更新日時'])
  attended = cur.fetchall()
  for row in attended:
    ws.append(row)

  wb.save(f"Attend_{ args.section_id }.xlsx")


def export_all():
  # WorkBookの作成
  wb = openpyxl.Workbook()
  sheet = wb.active
  wb.remove(sheet) # Sheet1の削除

  cur.execute("select section_id from attendance group by section_id order by section_id")
  sections = cur.fetchall()
  for section in sections:
    # Sheetの作成
    ws = wb.create_sheet()
    ws.title = str(section[0])

    cur.execute("select id, card_id, student_id, create_datetime, update_datetime from attendance where section_id = ? order by student_id", (section[0], ))
    attended = cur.fetchall()
    for row in attended:
      ws.append(row)

  wb.save("Attend_all.xlsx")


def main():

  if args.section_id == 0:
    export_all()
  else:
    export()

if __name__ == "__main__":
  try:
    main()
  finally:
    cur.close()
    con.close()

